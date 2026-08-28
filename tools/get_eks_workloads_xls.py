#!/usr/bin/env python3
"""Generate an Excel report from k8s-workloads inventory JSON.

Each namespace becomes a sheet. Each row is a deployment name mapped to
the EKS cluster(s) where it runs.

Usage:
    python generate_k8s_workloads_xls.py --input <path_to_json> [--output <dir>]
"""

import argparse
import json
import os
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font


def load_inventory(path: str) -> dict:
    with open(path) as f:
        return json.load(f)


def build_namespace_map(data: dict, exclude_clusters: set[str] | None = None) -> dict[str, dict[str, list[str]]]:
    """Returns {namespace: {deployment_name: [cluster1, cluster2, ...]}}."""
    ns_map: dict[str, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    exclude = exclude_clusters or set()

    for region, clusters in data.get("regions", {}).items():
        for cluster_name, cluster_data in clusters.items():
            if cluster_name in exclude:
                continue
            if cluster_data.get("status") != "ok":
                continue
            for ns, ns_data in cluster_data.get("namespaces", {}).items():
                for dep in ns_data.get("deployments", []):
                    dep_name = dep["name"]
                    if cluster_name not in ns_map[ns][dep_name]:
                        ns_map[ns][dep_name].append(cluster_name)

    return ns_map


def write_xls(ns_map: dict[str, dict[str, list[str]]], output_path: str) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    bold = Font(bold=True)

    for ns in sorted(ns_map.keys()):
        # Excel sheet names max 31 chars, no special chars
        sheet_title = ns[:31]
        ws = wb.create_sheet(title=sheet_title)
        ws.append(["Deployment Name", "EKS Clusters"])
        ws["A1"].font = bold
        ws["B1"].font = bold

        for dep_name in sorted(ns_map[ns].keys()):
            clusters_str = ", ".join(sorted(ns_map[ns][dep_name]))
            ws.append([dep_name, clusters_str])

        # auto-width columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)

    wb.save(output_path)
    print(f"Written: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Generate XLS from k8s-workloads inventory JSON")
    parser.add_argument("--input", required=True, help="Path to k8s-workloads JSON file")
    parser.add_argument("--output", default=None, help="Output directory (default: same directory as input file)")
    parser.add_argument("--except", dest="exclude_clusters", nargs="+", default=[],
                        help="EKS cluster names to exclude (e.g. --except scnx-dts01-eks)")
    args = parser.parse_args()

    # Default output to same directory as input (e.g. output/<account_id>/k8s-workloads/)
    output_dir = args.output if args.output else os.path.dirname(os.path.abspath(args.input))

    data = load_inventory(args.input)
    ns_map = build_namespace_map(data, exclude_clusters=set(args.exclude_clusters))

    os.makedirs(output_dir, exist_ok=True)
    base = os.path.splitext(os.path.basename(args.input))[0]
    output_path = os.path.join(output_dir, f"{base}.xlsx")

    write_xls(ns_map, output_path)


if __name__ == "__main__":
    main()
